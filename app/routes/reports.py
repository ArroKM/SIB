"""
Report routes.
Handles all report exports (Excel and PDF).
"""
import io
import os
from datetime import datetime

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required

from app.config import Config
from app.database import get_db
from app.utils.helpers import query_all, query_one, build_barang_query, flatten_barang

reports_bp = Blueprint('reports', __name__)


# ---------------------------------------------------------------------------
# Export Excel Report (all surat)
# ---------------------------------------------------------------------------

@reports_bp.route('/report/excel')
@login_required
def export_excel():
    """Export all surat as Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db()
    rows = query_all(conn, "SELECT * FROM surat_izin ORDER BY tanggal DESC")
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Surat Izin"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = f'LAPORAN SURAT IZIN KELUAR MASUK BARANG - {Config.COMPANY_NAME}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['No', 'Jenis', 'No. Surat', 'Tanggal', 'Divisi',
               'Nama', 'Perusahaan', 'Status', 'Dibuat']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, r in enumerate(rows, 5):
        vals = [i - 4, r['jenis'].upper(), r['no_surat'], str(r['tanggal']),
                r['divisi'], r['nama'], r['perusahaan'],
                r['status'].upper(), str(r['created_at'])]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col in (1, 2, 8) else 'left')

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=4):
            for cell in row_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'laporan_surat_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ---------------------------------------------------------------------------
# Export PDF Report (all surat)
# ---------------------------------------------------------------------------

@reports_bp.route('/report/pdf')
@login_required
def export_report_pdf():
    """Export all surat as PDF report."""
    conn = get_db()
    rows = query_all(conn, "SELECT * FROM surat_izin ORDER BY tanggal DESC")
    total_keluar = query_one(conn,
                             "SELECT COUNT(*) AS c FROM surat_izin WHERE jenis='keluar'")['c']
    total_masuk = query_one(conn,
                            "SELECT COUNT(*) AS c FROM surat_izin WHERE jenis='masuk'")['c']
    conn.close()

    html = render_template('report_pdf.html', rows=rows,
                           total_keluar=total_keluar, total_masuk=total_masuk,
                           now=datetime.now())
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html).write_pdf()
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf', as_attachment=True,
            download_name=f'laporan_{datetime.now().strftime("%Y%m%d")}.pdf',
        )
    except Exception as e:
        flash(f'Gagal membuat PDF: {e}', 'danger')
        return redirect(url_for('dashboard.dashboard'))


# ---------------------------------------------------------------------------
# Barang Report (with filtering)
# ---------------------------------------------------------------------------

@reports_bp.route('/report/barang')
@login_required
def report_barang():
    """Display barang (items) report with filtering."""
    conn = get_db()
    q, p = build_barang_query(request.args)
    rows = query_all(conn, q, p)
    conn.close()
    items = flatten_barang(rows)

    total_keluar = sum(1 for r in rows if r['jenis'] == 'keluar')
    total_masuk = sum(1 for r in rows if r['jenis'] == 'masuk')

    return render_template('report_barang.html', items=items,
                           total_keluar=total_keluar, total_masuk=total_masuk,
                           total_barang=len(items),
                           jenis=request.args.get('jenis', ''),
                           status=request.args.get('status', ''),
                           date_from=request.args.get('date_from', ''),
                           date_to=request.args.get('date_to', ''),
                           search=request.args.get('q', ''))


# ---------------------------------------------------------------------------
# Barang Excel Export
# ---------------------------------------------------------------------------

@reports_bp.route('/report/barang/excel')
@login_required
def report_barang_excel():
    """Export barang report as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter

    conn = get_db()
    q, p = build_barang_query(request.args)
    rows = query_all(conn, q, p)
    conn.close()
    items = flatten_barang(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Barang"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.merge_cells('A1:M1')
    ws['A1'] = f'LAPORAN BARANG MASUK & KELUAR - {Config.COMPANY_NAME}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    filters_text = f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    jenis_f = request.args.get('jenis', '')
    status_f = request.args.get('status', '')
    if jenis_f:
        filters_text += f' | Jenis: {jenis_f.upper()}'
    if status_f:
        filters_text += f' | Status: {status_f.upper()}'
    ws['A2'] = filters_text
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['No', 'Jenis', 'No. Surat', 'Tanggal', 'Divisi', 'Pemohon',
               'Perusahaan', 'Nama Barang', 'Jumlah', 'Satuan', 'Keterangan',
               'Cek User', 'Cek Satpam']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, item in enumerate(items, 5):
        user_approval_text = ''
        if item['approval_user'] == 'sesuai':
            user_approval_text = '✓ Sesuai'
        elif item['approval_user'] == 'tidak_sesuai':
            user_approval_text = '✗ Tidak Sesuai'

        approval_text = ''
        if item['approval_satpam'] == 'sesuai':
            approval_text = '✓ Sesuai'
        elif item['approval_satpam'] == 'tidak_sesuai':
            approval_text = '✗ Tidak Sesuai'

        vals = [i - 4, item['jenis'].upper(), item['no_surat'],
                str(item['tanggal']), item['divisi'], item['nama_pemohon'],
                item['perusahaan'], item['nama_barang'],
                item['jumlah'], item['satuan'],
                item['keterangan'] or '-', user_approval_text, approval_text]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col in (1, 2, 9, 10, 12, 13) else 'left')
            if col == 12 and item['approval_user'] == 'sesuai':
                cell.fill = green_fill
            elif col == 12 and item['approval_user'] == 'tidak_sesuai':
                cell.fill = red_fill
            if col == 13 and item['approval_satpam'] == 'sesuai':
                cell.fill = green_fill
            elif col == 13 and item['approval_satpam'] == 'tidak_sesuai':
                cell.fill = red_fill

        # Add photos in column N onward
        if item['foto']:
            foto_col = len(headers) + 1
            for foto_name in item['foto']:
                foto_path = os.path.join(Config.UPLOAD_DIR, foto_name)
                if os.path.exists(foto_path):
                    try:
                        img = XlImage(foto_path)
                        img.width = 60
                        img.height = 60
                        cell_ref = f'{get_column_letter(foto_col)}{i}'
                        ws.add_image(img, cell_ref)
                        ws.row_dimensions[i].height = 50
                        foto_col += 1
                    except Exception:
                        pass

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=4):
            for cell in row_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'laporan_barang_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ---------------------------------------------------------------------------
# Barang PDF Export
# ---------------------------------------------------------------------------

@reports_bp.route('/report/barang/pdf')
@login_required
def report_barang_pdf():
    """Export barang report as PDF."""
    conn = get_db()
    q, p = build_barang_query(request.args)
    rows = query_all(conn, q, p)
    conn.close()
    items = flatten_barang(rows)

    total_keluar = sum(1 for r in rows if r['jenis'] == 'keluar')
    total_masuk = sum(1 for r in rows if r['jenis'] == 'masuk')

    html = render_template('report_barang_pdf.html', items=items,
                           total_keluar=total_keluar, total_masuk=total_masuk,
                           total_barang=len(items),
                           upload_dir=os.path.abspath(Config.UPLOAD_DIR),
                           now=datetime.now())
    try:
        from weasyprint import HTML
        from pathlib import Path
        base = Path('static').as_uri() + '/'
        pdf_bytes = HTML(string=html, base_url=base).write_pdf()
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf', as_attachment=True,
            download_name=f'laporan_barang_{datetime.now().strftime("%Y%m%d")}.pdf',
        )
    except Exception as e:
        flash(f'Gagal membuat PDF: {e}', 'danger')
        return redirect(url_for('reports.report_barang'))
